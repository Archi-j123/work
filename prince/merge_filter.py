import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog
import os
from datetime import datetime

# Function to get file paths using a file dialog
def get_file_path():
    root = Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    return file_path

# Function to compare and merge two dataframes, highlighting differences
def compare_and_highlight(df1, df2, output_file):
    # Ensure both dataframes have the same columns
    all_columns = pd.concat([df1, df2], axis=1, keys=["File 1", "File 2"])

    # Fill missing values with 'N/A'
    all_columns = all_columns.fillna("N/A")
    
    # Create a new workbook and add a sheet
    wb = Workbook()
    ws = wb.active
    
    # Define a yellow fill for highlighting differences
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Write headers to the Excel sheet
    for col_num, col_name in enumerate(all_columns.columns, start=1):
        ws.cell(row=1, column=col_num).value = f"{col_name[0]} - {col_name[1]}"
    
    # Compare each cell and highlight differences
    max_rows = all_columns.shape[0]
    max_cols = all_columns.shape[1]

    for row in range(max_rows):
        for col in range(0, max_cols, 2):
            val1 = all_columns.iloc[row, col]
            val2 = all_columns.iloc[row, col+1] if col+1 < max_cols else None
            
            # Write values from both DataFrames to the sheet
            ws.cell(row=row+2, column=col+1).value = val1
            ws.cell(row=row+2, column=col+2).value = val2
            
            # If the values don't match, highlight both cells
            if val1 != val2:
                ws.cell(row=row+2, column=col+1).fill = yellow_fill
                ws.cell(row=row+2, column=col+2).fill = yellow_fill

    # Save the output file
    wb.save(output_file)
    print(f"Comparison complete. Differences highlighted in {output_file}")

# Main function to handle file selection and processing
def main():
    print("Select the first CSV file:")
    file1 = get_file_path()
    
    print("Select the second CSV file:")
    file2 = get_file_path()
    
    # Read the CSV files into DataFrames
    df1 = pd.read_csv(file1)
    df2 = pd.read_csv(file2)

    # print("First file preview:")
    # print(df1.head())
    # print("Second file preview:")
    # print(df2.head())
    
    # Dynamic output file name with timestamp to avoid overwriting
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"comparison_result_{timestamp}.xlsx"
    
    # Compare and highlight differences across all columns
    compare_and_highlight(df1, df2, output_file)

if __name__ == "__main__":
    main()
